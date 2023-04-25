import traceback
import openpyxl
from threshold_and_min_quality.Reqests import Reqests


class Get_all_lms_data(Reqests):
    def __init__(self) -> None:
        super().__init__()

    def creating_table(self):
        self.Wordbook = openpyxl.Workbook()
        self.sheet = self.Wordbook['Sheet']
        self.sheet.title = "Отчет"
        column_name = ["id ЛМ", "Название ЛМ", "Логин создателя", "Вендор", "Количество фото", "Порог Схожести", "Событий за неделю", "Среднее кол-во событий в день", "Минимальное качество лица в ЛМ"]
        for column, val in enumerate(column_name):
            self.sheet.cell(1, column+1).value = val


    def vif_completion(self):
        res = self.get_lm(22, self.lm_id )
        try:
            self.sheet.cell(self.sheet.max_row, 5).value = res.get("monitoring").get("faces", "Не удалось получить количество лиц")
        except:
            self.sheet.cell(self.sheet.max_row, 5).value = "Нет данных"
        try:
            self.sheet.cell(self.sheet.max_row, 6).value = res.get("monitoring").get("threshold")
        except:
            self.sheet.cell(self.sheet.max_row, 6).value = "Нет данных"
        min = 1
        faces = self.get_lm_faces(22, self.lm_id )
        try:
            if res.get("monitoring").get("faces", "Не удалось получить количество лиц") == 0:
                self.sheet.cell(self.sheet.max_row, 9).value = "Нет лиц"
            else:
                while True:
                    for face in faces.get("faces"):
                        count = 0
                        qal = 0
                        try:
                            if face.get("features").get("quality").get("ntechlab", "0") != "0":
                                qal += face.get("features").get("quality").get("ntechlab")
                                count += 1
                            if face.get("features").get("quality").get("synesis", "0") != "0":
                                qal += face.get("features").get("quality").get("synesis")
                                count += 1
                            if face.get("features").get("quality").get("tevian", "0") != "0":
                                qal += face.get("features").get("quality").get("tevian")
                                count += 1
                            if face.get("features").get("quality").get("visionlabs", "0") != "0":
                                qal += face.get("features").get("quality").get("visionlabs")
                                count += 1
                            if qal / count < min:
                                min = qal / count
                        except:
                            pass
                    if str(faces.get("next_page", 0)) != "0" and str(faces.get("next_page", 0)) != "None":
                        faces = self.get_lm_faces(22, self.lm_id , page=faces.get("next_page", 0))
                    else:
                        break
                self.sheet.cell(self.sheet.max_row, 9).value = min
        except:
            self.sheet.cell(self.sheet.max_row, 9).value = "Нет данных"


    def par_completion(self):
        try:
            res = self.get_qal_and_tr(self.lm_id )
            if res.get("conditionsType") == "faces:*":
                self.sheet.cell(self.sheet.max_row, 5).value = "Все лица"
                self.sheet.cell(self.sheet.max_row, 6).value = "Все лица"
                self.sheet.cell(self.sheet.max_row, 9).value = "Все лица"


            elif res.get("conditionsType") == "faces":
                count_faces = 0
                try:
                    count_faces = self.get_lm(25, self.lm_id ).get("monitoring").get("faces")
                    self.sheet.cell(self.sheet.max_row, 5).value = count_faces
                except:
                    self.sheet.cell(self.sheet.max_row, 5).value = "Нет данных"
                self.sheet.cell(self.sheet.max_row, 56).value = res.get("threshold")
                if count_faces == 0:
                    self.sheet.cell(self.sheet.max_row, 9).value = "Нет лиц"
                else:
                    faces = self.get_lm_faces(25, self.lm_id )
                    min = 1
                    while True:
                        try:
                            for i in faces.get("faces"):
                                try:
                                    if float(i.get("score", 1)) < min and float(i.get("score", 1)) > 0:
                                        min = i.get("score")
                                except:
                                    pass
                            if str(faces.get("next_page", 0)) != "0" and str(
                                    faces.get("next_page", 0)) != "None":
                                faces = self.get_lm_faces(25, self.lm_id , page=faces.get("next_page"))
                            else:
                                self.sheet.cell(self.sheet.max_row, 9).value = min
                                break

                        except:
                            self.sheet.cell(self.sheet.max_row, 9).value = min
                    try:
                        self.sheet.cell(self.sheet.max_row, 5).value = self.get_lm(25, self.lm_id ).get("monitoring").get("faces")
                    except:
                        self.sheet.cell(self.sheet.max_row, 5).value = "Нет данных"

            elif res.get("conditionsType") == "galleries":
                if res.get("conditions", 0) == 0:
                    self.sheet.cell(self.sheet.max_row, 5).value = "Нет лиц"
                    self.sheet.cell(self.sheet.max_row, 6).value = "Нет лиц"
                    self.sheet.cell(self.sheet.max_row, 9).value = "Нет лиц"
                else:
                    thresholds = ""
                    faces = 0
                    min = 1
                    for gal in res.get("conditions", 0):
                        thresholds = f"{thresholds}{gal.get('threshold')};"
                        faces += self.get_gal(25, gal.get('id')).get("gallery").get("faces")
                        qal_faces = self.get_gal_face(gal.get('id'))
                        while True:
                            for face in qal_faces.get("faces"):
                                if float(face.get("score", 1)) < min and float(face.get("score", 1)) > 0:
                                    min = face.get("score")
                            if str(qal_faces.get("next_page", 0)) != "0" and str(
                                    qal_faces.get("next_page", 0)) != "None":
                                qal_faces = self.get_gal_face(gal.get('id'), page=qal_faces.get("next_page"))
                            else:
                                break
                    self.sheet.cell(self.sheet.max_row, 5).value = faces
                    self.sheet.cell(self.sheet.max_row, 6).value = thresholds
                    self.sheet.cell(self.sheet.max_row, 9).value = min

        except Exception as ex:
            traceback.print_exc()
            self.sheet.cell(self.sheet.max_row, 5).value = "Нет данных"
            self.sheet.cell(self.sheet.max_row, 6).value = "Нет данных"
            self.sheet.cell(self.sheet.max_row, 9).value = "Нет данных"


    def main_function(self):
        self.creating_table()
        lms = self.get_all_lm()
        for index, lm in enumerate(lms.get('rules')):
            print(index)
            try:
                self.lm_id = lm.get("id")
                if str(lm.get("vendor")) != "20":
                    self.sheet.cell(self.sheet.max_row + 1, 1).value = lm.get("id")
                    self.sheet.cell(self.sheet.max_row, 2).value = lm.get("title")
                    self.sheet.cell(self.sheet.max_row, 3).value = lm.get("ownerName")
                    self.sheet.cell(self.sheet.max_row, 4).value = lm.get("vendor")

                if str(lm.get("vendor")) == "22":
                    try:
                        self.vif_completion()
                        events = self.events_count(22)
                        self.sheet.cell(self.sheet.max_row, 7).value = events[0]
                        self.sheet.cell(self.sheet.max_row, 8).value = events[1]
                    except Exception as ex:
                        traceback.print_exc()
                        self.sheet.cell(self.sheet.max_row, 2).value = "Ошибка запроса"
                        self.sheet.cell(self.sheet.max_row, 3).value = "Ошибка запроса"
                        self.sheet.cell(self.sheet.max_row, 4).value = "Ошибка запроса"

                if str(lm.get("vendor")) == "25":
                    self.par_completion()
                    events = self.events_count(25)
                    self.sheet.cell(self.sheet.max_row, 7).value = events[0]
                    self.sheet.cell(self.sheet.max_row, 8).value = events[1]

            except Exception as ex:
                print("Критическа ошибка")
        self.Wordbook.save(f'Детекты.xlsx')


    def events_count(self, vender):
        count_days = []
        count_events = 0
        try:
            next_page = "null"
            while True:
                events = self.get_events(vender, self.lm_id, next=next_page)
                for event in events.get("events"):
                    try:
                        if event.get("face").get("timestamp").split("T")[0] not in count_days:
                            count_days.append(event.get("face").get("timestamp").split("T")[0])
                        count_events += 1
                    except Exception as ex:
                        traceback.print_exc()
                        print(event)

                if str(events.get("next_page")) == next_page or str(events.get("next_page", "None")) == "None":
                    break
                next_page = events.get("next_page")
            if len(count_days) != 0:
                return count_events, count_events / len(count_days)
            else:
                return 0, 0
        except Exception as ex:
            traceback.print_exc()
            try:
                return count_events, count_events / len(count_days)
            except:
                return "Нет данных", "Нет данных"

